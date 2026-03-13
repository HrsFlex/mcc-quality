"""
weekly_summary.py — Generates a branded Excel report from the quality_incidents table.
Produces the same sheets as analyze_insights.py but sourced from the DB.
"""
import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _style_header_row(ws, header_fill_hex="003366"):
    fill = PatternFill(start_color=header_fill_hex, end_color=header_fill_hex, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_col_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def generate(conn, output_dir: str, run_id: str) -> str:
    """
    Pull data from quality_incidents, build aggregated views, write Excel.
    Returns the output file path.
    """
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = os.path.join(output_dir, f"Quality_Weekly_Summary_{timestamp}.xlsx")

    logger.info("[Report] Querying data for weekly summary...")

    df = pd.read_sql_query(
        "SELECT * FROM quality_incidents",
        conn,
        parse_dates=["incident_date"],
    )

    if df.empty:
        logger.warning("[Report] No data in quality_incidents — skipping report generation")
        return ""

    # ── Aggregations ──────────────────────────────────────────────────────────────
    df["Quarter"] = df["incident_date"].dt.to_period("Q").astype(str)

    time_agg = df.groupby("source").agg(
        Earliest_Date=("incident_date", "min"),
        Latest_Date=("incident_date", "max"),
    ).reset_index()

    source_agg = df.groupby("source").agg(
        Incidents=("source", "count"),
        Total_Credit_USD=("credit_requested_usd", "sum"),
    ).reset_index().sort_values("Total_Credit_USD", ascending=False)

    quarterly_pivot = df.pivot_table(
        index="Quarter",
        columns="source",
        values="credit_requested_usd",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    plant_agg = df.groupby("plant").agg(
        Incidents=("plant", "count"),
        Total_Credit_USD=("credit_requested_usd", "sum"),
    ).reset_index().sort_values("Total_Credit_USD", ascending=False)

    defect_agg = df.groupby("defect_category").agg(
        Incidents=("defect_category", "count"),
        Total_Credit_USD=("credit_requested_usd", "sum"),
    ).reset_index().sort_values("Total_Credit_USD", ascending=False).head(15)

    customer_agg = df.groupby("customer").agg(
        Incidents=("customer", "count"),
        Total_Credit_USD=("credit_requested_usd", "sum"),
    ).reset_index().sort_values("Total_Credit_USD", ascending=False).head(20)

    cust_plant = df.groupby(["customer", "plant"]).agg(
        Total_Credit_USD=("credit_requested_usd", "sum")
    ).reset_index().sort_values("Total_Credit_USD", ascending=False).head(15)

    plant_defect = df.groupby(["plant", "defect_category"]).agg(
        Total_Credit_USD=("credit_requested_usd", "sum")
    ).reset_index().sort_values("Total_Credit_USD", ascending=False).head(15)

    # Summary stats row
    summary = pd.DataFrame([{
        "Pipeline Run ID": run_id,
        "Generated At": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Total Incidents": len(df),
        "Total Credit (USD)": f"${df['credit_requested_usd'].sum():,.2f}",
        "Sources": ", ".join(df["source"].unique().tolist()),
        "Plants": df["plant"].nunique(),
        "Customers": df["customer"].nunique(),
    }])

    # ── Write Excel ───────────────────────────────────────────────────────────────
    sheets = {
        "Summary":             summary,
        "By Source":           source_agg,
        "Quarterly Breakdown": quarterly_pivot,
        "By Plant":            plant_agg,
        "By Defect Category":  defect_agg,
        "By Customer":         customer_agg,
        "Time Periods":        time_agg,
        "Customer x Plant":    cust_plant,
        "Plant x Defect":      plant_defect,
        "Raw Data":            df.drop(columns=["id"], errors="ignore"),
    }

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, data in sheets.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _style_header_row(ws)
            _auto_col_width(ws)

    logger.info(f"[Report] Weekly summary saved → {out_path}")
    return out_path
